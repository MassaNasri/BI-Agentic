"""
Enhanced Extractor Service Kafka Listener
Extracts data from files/databases and emits comprehensive metadata
"""
import os
import sys
import logging
import threading
import time
import hashlib
import json
from collections import deque
import pandas as pd
from datetime import datetime, timezone
from typing import Dict, Any, List
from uuid import uuid4
from clickhouse_driver import Client
from shared.utils.kafka_consumer import KafkaMessageConsumer
from shared.utils.kafka_producer import KafkaMessageProducer
from shared.utils.metadata_schema import MetadataSchema
from shared.utils.credential_encryption import get_encryption_instance
from shared.utils.idempotency_manager import (
    IdempotencyClaim,
    IdempotencyManager,
    IdempotencyKey,
    PipelineStage,
)
from shared.utils.logging_utils import configure_logging
from shared.utils.metrics import ROWS_PROCESSED, ERRORS_TOTAL, PROCESS_LATENCY, start_metrics_server
from shared.utils.health_server import start_health_server
from shared.utils.lineage_tracker import LineageTracker
from shared.utils.db_type_utils import normalize_db_type
from shared.models.lineage import LineageRecord
from shared.utils.tracing import configure_tracing, get_tracer, set_trace_id_from_span
from shared.utils.schema_contract_store import build_schema_contract_store_from_env
from .db_connector import DBConnector
from .row_extractor import RowExtractor
from .csv_extraction_strategy import CSVExtractionStrategy
from .database_extraction_strategy import DatabaseExtractionStrategy
from .extraction_strategy import ExtractionConfig

logger = logging.getLogger(__name__)

configure_logging()
configure_tracing("extractor-service")
start_metrics_server(int(os.getenv("EXTRACTOR_METRICS_PORT", "9101")))
start_health_server(int(os.getenv("EXTRACTOR_HEALTH_PORT", "8083")))


class ConnectionListener:
    """
    Listens to connection_topic and extracts data from files or databases
    """

    def __init__(self):
        self.consumer = KafkaMessageConsumer("connection_topic")
        self.producer = KafkaMessageProducer()
        self.db_connector = DBConnector()
        self.row_extractor = RowExtractor()
        self.csv_strategy = CSVExtractionStrategy()
        self.database_strategy = DatabaseExtractionStrategy()
        try:
            self.lineage_tracker = LineageTracker()
        except Exception as exc:
            logger.warning("[EXTRACTOR] Lineage tracker unavailable, continuing without lineage side-effects: %s", exc)
            self.lineage_tracker = None
        
        # Initialize ClickHouse client for idempotency checks
        try:
            self.clickhouse_client = Client(
                host=os.getenv('CLICKHOUSE_HOST', 'clickhouse'),
                port=int(os.getenv('CLICKHOUSE_PORT', '9000')),
                user=os.getenv('CLICKHOUSE_USER', 'default'),
                password=os.getenv('CLICKHOUSE_PASSWORD', ''),
                database=os.getenv('CLICKHOUSE_DATABASE', 'etl')
            )
            logger.info("[EXTRACTOR] ClickHouse client initialized for idempotency checks")
        except Exception as e:
            logger.warning(f"[EXTRACTOR] Failed to initialize ClickHouse client: {e}")
            self.clickhouse_client = None
        
        # Initialize IdempotencyManager with ClickHouse client
        self.idempotency_manager = IdempotencyManager(self.clickhouse_client) if self.clickhouse_client else None
        self.schema_contract_store = build_schema_contract_store_from_env(self.clickhouse_client)
        
        if self.idempotency_manager:
            logger.info("[EXTRACTOR] IdempotencyManager initialized successfully")
        else:
            logger.warning("[EXTRACTOR] IdempotencyManager not available - idempotency checks disabled")
        self.batch_size = int(os.getenv("EXTRACTOR_BATCH_SIZE", "500"))
        self.max_error_entries = int(os.getenv("EXTRACTOR_MAX_ERROR_ENTRIES", "100"))
        self.send_retries = int(os.getenv("EXTRACTOR_SEND_RETRIES", "3"))
        self.send_backoff_base = float(os.getenv("EXTRACTOR_SEND_BACKOFF_BASE", "0.5"))
        self.send_backoff_max = float(os.getenv("EXTRACTOR_SEND_BACKOFF_MAX", "5.0"))
        self.dlq_topic = os.getenv("EXTRACTOR_DLQ_TOPIC", "extracted_rows_dlq")
        self._db_last_pk_state: Dict[str, Any] = {}
        self._db_state_lock = threading.Lock()

    def _send_with_retry(self, topic: str, message: Dict[str, Any], context: str) -> bool:
        for attempt in range(1, self.send_retries + 1):
            if self.producer.send(topic, message):
                return True
            if attempt < self.send_retries:
                backoff = min(self.send_backoff_max, self.send_backoff_base * (2 ** (attempt - 1)))
                logger.warning(
                    "[EXTRACTOR] Send failed for %s (attempt %s/%s), retrying in %.2fs",
                    context,
                    attempt,
                    self.send_retries,
                    backoff,
                )
                time.sleep(backoff)
        logger.error("[EXTRACTOR] Exhausted retries for %s", context)
        return False

    def _send_to_dlq(self, reason: str, payload: Dict[str, Any]) -> bool:
        dlq_message = {
            "stage": "extract",
            "reason": reason,
            "failed_at": datetime.utcnow().isoformat(),
            "payload": payload,
        }
        return self.producer.send(self.dlq_topic, dlq_message, validate=False)

    def _send_batch(self, base_message: Dict[str, Any], rows: List[Dict[str, Any]]) -> bool:
        if not rows:
            return True
        source = base_message.get("source") or base_message.get("source_id") or "unknown"
        batch_message = {
            **base_message,
            "source": source,
            "source_id": source,
            "schema_version": base_message.get("schema_version") or "derived_unknown",
            "rows": rows,
            "row_count": len(rows),
        }
        if self._send_with_retry("extracted_rows_topic", batch_message, "extracted_rows_topic batch"):
            return True
        return self._send_to_dlq("kafka_send_failed", batch_message)

    def _safe_record_lineage(self, record: LineageRecord) -> None:
        """
        Lineage side-effects must not fail extraction.
        """
        if not self.lineage_tracker:
            return
        try:
            self.lineage_tracker.record_transformation(record)
        except Exception as exc:
            logger.warning("[EXTRACTOR] Lineage write failed and was skipped: %s", exc)

    def _derive_schema_version(self, schema_payload: Dict[str, Any]) -> str:
        normalized = json.dumps(schema_payload, sort_keys=True, default=str)
        return f"sv_{hashlib.sha256(normalized.encode('utf-8')).hexdigest()[:12]}"

    def _resolve_schema_contract(self, source_id: str, schema_version: str):
        try:
            contract = self.schema_contract_store.get_contract(source_id, schema_version)
            return contract.to_dict() if contract else None
        except Exception as exc:
            logger.warning(
                "[EXTRACTOR] Failed schema contract lookup for source=%s version=%s: %s",
                source_id,
                schema_version,
                exc,
            )
            return None

    def _build_idempotency_key(self, source_id: str, row_hash: str) -> IdempotencyKey:
        # Use stage-scoped batch identifier so retries of the same source row hash are deduplicated.
        return IdempotencyKey(
            source_id=source_id,
            batch_id=f"extract:{source_id}",
            row_hash=row_hash,
        )

    def process_file(self, message: Dict[str, Any]) -> bool:
        """
        Extract schema and rows from a file with metadata emission.
        Implements idempotency checks to prevent duplicate extraction.
        """
        source_id = message.get("filename", "unknown")
        file_path = message.get("path")
        batch_id = str(uuid4())

        try:
            start_time = datetime.now()
            logger.info("[EXTRACTOR] Processing file: %s (batch_id: %s)", source_id, batch_id)
            if not file_path:
                logger.error("[EXTRACTOR ERROR] Missing file path")
                return False

            schema = {"source": source_id, "type": "file", "columns": [], "dtypes": {}, "row_count": 0}
            if file_path.endswith(".csv"):
                head_df = pd.read_csv(file_path, nrows=1)
                schema["columns"] = list(head_df.columns)
                schema["dtypes"] = {col: str(dtype) for col, dtype in head_df.dtypes.items()}

                csv_config = ExtractionConfig(
                    source_id=source_id,
                    source_type="csv",
                    connection_params={"file_path": file_path},
                    batch_size=self.batch_size,
                )
                file_iterator = self.csv_strategy.iter_batches(csv_config, batch_size=self.batch_size)
            elif file_path.endswith((".xls", ".xlsx")):
                try:
                    from openpyxl import load_workbook
                except Exception as exc:
                    logger.error("[EXTRACTOR ERROR] openpyxl is required for Excel extraction: %s", exc)
                    return False
                workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                worksheet = workbook.active
                rows_iter = worksheet.iter_rows(values_only=True)
                headers = next(rows_iter, None) or []
                schema["columns"] = [str(col) if col is not None else "" for col in headers]

                def excel_chunks():
                    batch = []
                    for values in rows_iter:
                        row_dict = {}
                        for idx, column in enumerate(schema["columns"]):
                            key = column if column else f"column_{idx}"
                            row_dict[key] = values[idx] if idx < len(values) else None
                        batch.append(row_dict)
                        if len(batch) >= self.batch_size:
                            yield batch
                            batch = []
                    if batch:
                        yield batch
                    workbook.close()

                file_iterator = excel_chunks()
            else:
                logger.error("[EXTRACTOR ERROR] Unsupported file type: %s", file_path)
                return False

            schema_version = message.get("schema_version") or self._derive_schema_version(schema)
            schema["version"] = schema_version
            schema["source_id"] = source_id
            schema_contract = self._resolve_schema_contract(source_id, schema_version)

            self._send_with_retry("schema_topic", schema, "schema_topic file schema")
            schema_metadata = MetadataSchema.create_schema_metadata(source_id=source_id, schema=schema, row_count=0)
            self._send_with_retry("metadata_topic", schema_metadata, "metadata_topic file schema")

            rows_successful = 0
            rows_failed = 0
            rows_skipped = 0
            rows_processed = 0
            row_id = 0
            error_count = 0
            errors = deque(maxlen=self.max_error_entries)
            batch_rows: List[Dict[str, Any]] = []
            batch_claims: List[IdempotencyClaim] = []
            send_failures = 0

            for chunk in file_iterator:
                row_items = chunk.to_dict("records") if isinstance(chunk, pd.DataFrame) else chunk
                prepared_rows: List[Dict[str, Any]] = []
                prepared_keys: List[IdempotencyKey] = []
                for row_dict in row_items:
                    current_row_id = row_id
                    row_id += 1
                    try:
                        rows_processed += 1
                        if self.idempotency_manager:
                            row_hash = self.idempotency_manager.generate_row_hash(row_dict)
                        else:
                            import hashlib
                            row_hash = hashlib.sha256(str(sorted(row_dict.items())).encode("utf-8")).hexdigest()

                        prepared_rows.append({
                            "row_id": current_row_id,
                            "row_dict": row_dict,
                            "row_hash": row_hash,
                        })
                        prepared_keys.append(
                            self._build_idempotency_key(source_id, row_hash)
                        )
                    except Exception as e:
                        rows_failed += 1
                        error_count += 1
                        errors.append(f"Row {current_row_id}: {str(e)}")

                allowed_keys = {k.to_dedup_key() for k in prepared_keys}
                claim_by_dedup: Dict[str, IdempotencyClaim] = {}
                if self.idempotency_manager and prepared_keys:
                    claims = self.idempotency_manager.claim_new_keys(prepared_keys, PipelineStage.EXTRACT)
                    claim_by_dedup = {claim.key.to_dedup_key(): claim for claim in claims}
                    allowed_keys = set(claim_by_dedup.keys())

                for row_meta, idempotency_key in zip(prepared_rows, prepared_keys):
                    if idempotency_key.to_dedup_key() not in allowed_keys:
                        rows_skipped += 1
                        continue

                    row_data = {
                        "source": source_id,
                        "source_id": source_id,
                        "batch_id": batch_id,
                        "row_id": int(row_meta["row_id"]),
                        "data": row_meta["row_dict"],
                        "_dedup_key": row_meta["row_hash"],
                        "_extracted_at": datetime.now(timezone.utc).isoformat(),
                    }
                    lineage_row_id = LineageTracker.deterministic_row_id(
                        source_id=source_id,
                        batch_id=batch_id,
                        dedup_key=row_meta["row_hash"],
                        stage=PipelineStage.EXTRACT.value,
                    )
                    row_data["_lineage_row_id"] = str(lineage_row_id)
                    self._safe_record_lineage(
                        LineageRecord(
                            row_id=lineage_row_id,
                            source_id=source_id,
                            batch_id=batch_id,
                            stage=PipelineStage.EXTRACT.value,
                            applied_rules=[],
                            parent_row_ids=[],
                        )
                    )
                    batch_rows.append(row_data)
                    if claim_by_dedup:
                        claim = claim_by_dedup.get(idempotency_key.to_dedup_key())
                        if claim:
                            batch_claims.append(claim)

                    if len(batch_rows) >= self.batch_size:
                        sent = self._send_batch(
                            {
                                "source": source_id,
                                "source_id": source_id,
                                "batch_id": batch_id,
                                "schema_version": schema_version,
                                "schema_contract": schema_contract,
                            },
                            batch_rows,
                        )
                        if sent:
                            rows_successful += len(batch_rows)
                        else:
                            rows_failed += len(batch_rows)
                            error_count += 1
                            errors.append("Failed to send batch")
                            send_failures += 1
                            if self.idempotency_manager and batch_claims:
                                self.idempotency_manager.rollback_claims(batch_claims, PipelineStage.EXTRACT)
                        batch_rows = []
                        batch_claims = []

            if batch_rows:
                sent = self._send_batch(
                    {
                        "source": source_id,
                        "source_id": source_id,
                        "batch_id": batch_id,
                        "schema_version": schema_version,
                        "schema_contract": schema_contract,
                    },
                    batch_rows,
                )
                if sent:
                    rows_successful += len(batch_rows)
                else:
                    rows_failed += len(batch_rows)
                    error_count += 1
                    errors.append("Failed to send final batch")
                    send_failures += 1
                    if self.idempotency_manager and batch_claims:
                        self.idempotency_manager.rollback_claims(batch_claims, PipelineStage.EXTRACT)

            extraction_metadata = MetadataSchema.create_extraction_metadata(
                source_id=source_id,
                rows_processed=rows_processed,
                rows_successful=rows_successful,
                rows_failed=rows_failed,
                errors=list(errors),
            )
            extraction_metadata["error_count"] = error_count
            extraction_metadata["error_sample_count"] = len(errors)
            self._send_with_retry("metadata_topic", extraction_metadata, "metadata_topic extraction")
            if PROCESS_LATENCY:
                PROCESS_LATENCY.labels(service="extractor", stage="extract").observe(
                    (datetime.now() - start_time).total_seconds()
                )
            logger.info(
                "[EXTRACTOR] Published %s rows (failed: %s, skipped duplicates: %s)",
                rows_successful,
                rows_failed,
                rows_skipped,
            )
            return send_failures == 0 and rows_failed == 0
        except Exception as e:
            if ERRORS_TOTAL:
                ERRORS_TOTAL.labels(service="extractor", stage="extract").inc()
            logger.error("[EXTRACTOR ERROR] Failed to process file %s: %s", source_id, e)
            return False

    def process_database(self, message: Dict[str, Any]) -> bool:
        """
        Extract schema and rows from a database with metadata emission.
        Implements idempotency checks to prevent duplicate extraction.
        
        Args:
            message: Message from connection_topic
        """
        batch_id = str(uuid4())  # Generate batch ID for this extraction
        
        connection = None
        try:
            start_time = datetime.now()
            # Decrypt password if encrypted
            encryption = get_encryption_instance()
            password = message["password"]
            
            if message.get("_password_encrypted", False):
                logger.info("[EXTRACTOR] Decrypting database password")
                password = encryption.decrypt(password)
            
            db_type = normalize_db_type(message.get("db_type"))
            if db_type is None:
                logger.error("[EXTRACTOR ERROR] Unsupported database type: %s", message.get("db_type"))
                return False

            db_config = {
                "db_type": db_type,
                "host": message["host"],
                "user": message["user"],
                "password": password,
                "database": message["database"],
                "port": message["port"]
            }
            
            source_id = f"{db_config['database']}"
            logger.info(f"[EXTRACTOR] Connecting to {db_config['db_type']} database: {source_id} (batch_id: {batch_id})")
            
            # Connect to database
            connection = self.db_connector.connect(db_config)
            cursor = connection.cursor()
            
            # Get list of tables
            if db_config["db_type"] == "mysql":
                cursor.execute("SHOW TABLES")
                tables = [row[0] for row in cursor.fetchall()]
            elif db_config["db_type"] == "postgres":
                cursor.execute("SELECT tablename FROM pg_tables WHERE schemaname='public'")
                tables = [row[0] for row in cursor.fetchall()]
            else:
                logger.error(f"[EXTRACTOR ERROR] Unsupported database type")
                return False
            
            logger.info(f"[EXTRACTOR] Found {len(tables)} tables")
            table_schema_versions: Dict[str, str] = {}
            table_schema_contracts: Dict[str, Any] = {}
            
            # Extract schema for each table
            for table in tables:
                # Use proper identifier quoting to prevent SQL injection
                # Table names come from database metadata queries but should still be quoted
                if db_config["db_type"] == "mysql":
                    # MySQL uses backticks for identifier quoting
                    quoted_table = f"`{table.replace('`', '``')}`"
                elif db_config["db_type"] == "postgres":
                    # PostgreSQL uses double quotes for identifier quoting
                    safe_table = table.replace('"', '""')
                    quoted_table = f'"{safe_table}"'
                else:
                    # Fallback: validate table name contains only safe characters
                    import re
                    if not re.match(r'^[a-zA-Z0-9_]+$', table):
                        logger.error(f"[EXTRACTOR ERROR] Invalid table name: {table}")
                        continue
                    quoted_table = table
                
                query = f"SELECT * FROM {quoted_table} LIMIT 1"
                cursor.execute(query)
                columns = [desc[0] for desc in cursor.description]
                
                table_source = f"{db_config['database']}.{table}"
                schema = {
                    "source": table_source,
                    "source_id": table_source,
                    "type": "database",
                    "table": table,
                    "columns": columns,
                    "db_type": db_config["db_type"]
                }
                schema_version = message.get("schema_version") or self._derive_schema_version(schema)
                schema["version"] = schema_version
                table_schema_versions[table_source] = schema_version
                table_schema_contracts[table_source] = self._resolve_schema_contract(table_source, schema_version)
                
                self._send_with_retry("schema_topic", schema, "schema_topic db schema")
                logger.info(f"[EXTRACTOR] Published schema for {table}: {len(columns)} columns")
                
                # Emit schema metadata
                schema_metadata = MetadataSchema.create_schema_metadata(
                    source_id=table_source,
                    schema=schema,
                    row_count=0  # Will be updated after extraction
                )
                self._send_with_retry("metadata_topic", schema_metadata, "metadata_topic db schema")
            
            # Extract rows with idempotency checks.
            total_rows = 0
            rows_successful = 0
            rows_failed = 0
            rows_skipped = 0  # Track skipped duplicates
            send_failures = 0
            error_count = 0
            errors = deque(maxlen=self.max_error_entries)
            order_by_config = message.get("order_by")

            for table in tables:
                table_source = f"{db_config['database']}.{table}"
                table_batch_rows: List[Dict[str, Any]] = []
                table_batch_claims: List[IdempotencyClaim] = []

                if isinstance(order_by_config, dict):
                    table_order_by = order_by_config.get(table)
                else:
                    table_order_by = order_by_config

                pk_column = self.database_strategy.detect_primary_key(connection, table)

                offset = 0
                with self._db_state_lock:
                    last_pk = self._db_last_pk_state.get(table_source)
                while True:
                    strategy_config = ExtractionConfig(
                        source_id=table_source,
                        source_type="database",
                        connection_params={
                            "connection": connection,
                            "table": table,
                            "order_by": table_order_by,
                            "pk_column": pk_column,
                            "last_pk": last_pk,
                        },
                        batch_size=self.batch_size,
                    )
                    batch = self.database_strategy.extract_batch(
                        strategy_config, offset=offset, limit=self.batch_size
                    )
                    if not batch.rows:
                        break

                    pagination_mode = batch.metadata.get("pagination_mode")
                    if pagination_mode == "keyset":
                        last_pk = batch.metadata.get("next_last_pk", last_pk)

                    prepared_rows: List[Dict[str, Any]] = []
                    prepared_keys: List[IdempotencyKey] = []
                    for row in batch.rows:
                        try:
                            row_dict = {k: v for k, v in row.items() if not str(k).startswith("_")}
                            if self.idempotency_manager:
                                row_hash = self.idempotency_manager.generate_row_hash(row_dict)
                            else:
                                import hashlib
                                sorted_items = sorted(row_dict.items())
                                row_hash = hashlib.sha256(str(sorted_items).encode('utf-8')).hexdigest()
                            prepared_rows.append({"row_dict": row_dict, "row_hash": row_hash})
                            prepared_keys.append(
                                self._build_idempotency_key(table_source, row_hash)
                            )
                        except Exception as e:
                            if ERRORS_TOTAL:
                                ERRORS_TOTAL.labels(service="extractor", stage="extract").inc()
                            rows_failed += 1
                            error_count += 1
                            errors.append(f"Table {table}: {str(e)}")
                            logger.warning(f"[EXTRACTOR] Error preparing row from {table}: {e}")

                    allowed_keys = {k.to_dedup_key() for k in prepared_keys}
                    claim_by_dedup: Dict[str, IdempotencyClaim] = {}
                    if self.idempotency_manager and prepared_keys:
                        claims = self.idempotency_manager.claim_new_keys(prepared_keys, PipelineStage.EXTRACT)
                        claim_by_dedup = {claim.key.to_dedup_key(): claim for claim in claims}
                        allowed_keys = set(claim_by_dedup.keys())

                    for row_meta, idempotency_key in zip(prepared_rows, prepared_keys):
                        if idempotency_key.to_dedup_key() not in allowed_keys:
                            rows_skipped += 1
                            continue

                        row_data = {
                            "source": table_source,
                            "source_id": table_source,
                            "batch_id": batch_id,
                            "table": table,
                            "data": row_meta["row_dict"],
                            "_dedup_key": row_meta["row_hash"],
                            "_extracted_at": datetime.now(timezone.utc).isoformat(),
                        }

                        lineage_row_id = LineageTracker.deterministic_row_id(
                            source_id=table_source,
                            batch_id=batch_id,
                            dedup_key=row_meta["row_hash"],
                            stage=PipelineStage.EXTRACT.value,
                        )
                        row_data["_lineage_row_id"] = str(lineage_row_id)
                        self._safe_record_lineage(
                            LineageRecord(
                                row_id=lineage_row_id,
                                source_id=table_source,
                                batch_id=batch_id,
                                stage=PipelineStage.EXTRACT.value,
                                applied_rules=[],
                                parent_row_ids=[],
                            )
                        )
                        table_batch_rows.append(row_data)
                        if claim_by_dedup:
                            claim = claim_by_dedup.get(idempotency_key.to_dedup_key())
                            if claim:
                                table_batch_claims.append(claim)
                        total_rows += 1

                        if len(table_batch_rows) >= self.batch_size:
                            sent = self._send_batch(
                                {
                                    "source": table_source,
                                    "source_id": table_source,
                                    "batch_id": batch_id,
                                    "schema_version": table_schema_versions.get(table_source, "derived_unknown"),
                                    "schema_contract": table_schema_contracts.get(table_source),
                                    "table": table,
                                },
                                table_batch_rows,
                            )
                            if sent:
                                rows_successful += len(table_batch_rows)
                                if ROWS_PROCESSED:
                                    ROWS_PROCESSED.labels(service="extractor", stage="extract").inc(len(table_batch_rows))
                            else:
                                if ERRORS_TOTAL:
                                    ERRORS_TOTAL.labels(service="extractor", stage="extract").inc(len(table_batch_rows))
                                rows_failed += len(table_batch_rows)
                                error_count += 1
                                errors.append(f"Table {table}: Failed to send batch")
                                send_failures += 1
                                if self.idempotency_manager and table_batch_claims:
                                    self.idempotency_manager.rollback_claims(table_batch_claims, PipelineStage.EXTRACT)
                            table_batch_rows = []
                            table_batch_claims = []
                    if not batch.has_more:
                        break
                    if pagination_mode == "keyset":
                        with self._db_state_lock:
                            self._db_last_pk_state[table_source] = last_pk
                    else:
                        offset += batch.total_rows

                # Flush remainder at the end of each table to avoid mixing table rows.
                if table_batch_rows:
                    sent = self._send_batch(
                        {
                            "source": table_source,
                            "source_id": table_source,
                            "batch_id": batch_id,
                            "schema_version": table_schema_versions.get(table_source, "derived_unknown"),
                            "schema_contract": table_schema_contracts.get(table_source),
                            "table": table,
                        },
                        table_batch_rows,
                    )
                    if sent:
                        rows_successful += len(table_batch_rows)
                        if ROWS_PROCESSED:
                            ROWS_PROCESSED.labels(service="extractor", stage="extract").inc(len(table_batch_rows))
                    else:
                        if ERRORS_TOTAL:
                            ERRORS_TOTAL.labels(service="extractor", stage="extract").inc(len(table_batch_rows))
                        rows_failed += len(table_batch_rows)
                        error_count += 1
                        errors.append(f"Table {table}: Failed to send final batch")
                        send_failures += 1
                        if self.idempotency_manager and table_batch_claims:
                            self.idempotency_manager.rollback_claims(table_batch_claims, PipelineStage.EXTRACT)

                if pk_column and last_pk is not None:
                    with self._db_state_lock:
                        self._db_last_pk_state[table_source] = last_pk

            logger.info(
                f"[EXTRACTOR] Published {rows_successful} rows "
                f"(failed: {rows_failed}, skipped duplicates: {rows_skipped})"
            )
            
            # Emit extraction metadata
            extraction_metadata = MetadataSchema.create_extraction_metadata(
                source_id=source_id,
                rows_processed=total_rows,
                rows_successful=rows_successful,
                rows_failed=rows_failed,
                errors=list(errors)
            )
            extraction_metadata["error_count"] = error_count
            extraction_metadata["error_sample_count"] = len(errors)
            self._send_with_retry("metadata_topic", extraction_metadata, "metadata_topic extraction")
            if PROCESS_LATENCY:
                duration = (datetime.now() - start_time).total_seconds()
                PROCESS_LATENCY.labels(service="extractor", stage="extract").observe(duration)
            return send_failures == 0 and rows_failed == 0
            
        except Exception as e:
            if ERRORS_TOTAL:
                ERRORS_TOTAL.labels(service="extractor", stage="extract").inc()
            logger.error(f"[EXTRACTOR ERROR] Failed to process database: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
        finally:
            if connection:
                connection.close()

    def listen(self):
        """Listen to connection_topic and process messages"""
        logger.info("[EXTRACTOR] Listening to connection_topic...")
        try:
            for message, record in self.consumer.listen_committable():
                logger.info(f"[EXTRACTOR] Received message: type={message.get('type')}")
                
                # Emit connection metadata
                connection_metadata = MetadataSchema.create_connection_metadata(
                    source_type=message.get("type", "unknown"),
                    source_id=message.get("filename") or message.get("database", "unknown"),
                    connection_info=message
                )
                self._send_with_retry("metadata_topic", connection_metadata, "metadata_topic connection")
                
                if message.get("type") == "file":
                    processed = self.process_file(message)
                elif message.get("type") == "database":
                    processed = self.process_database(message)
                else:
                    logger.error(f"[EXTRACTOR ERROR] Unknown message type: {message.get('type')}")
                    processed = False

                if processed:
                    self.consumer.commit(record)
                else:
                    logger.warning("[EXTRACTOR] Message processing failed, offset not committed")
        except KeyboardInterrupt:
            logger.info("[EXTRACTOR] Shutting down...")
        except Exception as e:
            logger.error(f"[EXTRACTOR] Fatal error: {e}")
            import traceback
            logger.error(traceback.format_exc())


def start_listener():
    """Entry point for the Kafka listener"""
    logger.info("[EXTRACTOR] Starting connection listener...")
    parallelism = int(os.getenv("KAFKA_CONSUMER_PARALLELISM", "1"))
    if parallelism <= 1:
        listener = ConnectionListener()
        listener.listen()
        return

    threads = []
    for i in range(parallelism):
        t = threading.Thread(
            target=lambda: ConnectionListener().listen(),
            name=f"extractor-consumer-{i+1}",
            daemon=True,
        )
        t.start()
        threads.append(t)

    for t in threads:
        t.join()
